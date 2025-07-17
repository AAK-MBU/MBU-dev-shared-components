"""
Integration tests for verifying Microsoft Office Excel functionality through SharePoint.

These tests validate that:
1. Authentication to SharePoint works using environment credentials.
2. Files in a SharePoint folder can be listed successfully.
3. A new row can be appended to an existing Excel file hosted on SharePoint.
4. The Excel sheet is correctly formatted and sorted after modification.

Env-vars required:
    SvcRpaMBU002_USERNAME
    SvcRpaMBU002_PASSWORD
"""

# pylint: disable=W0102, R0912, W0621


import os

import datetime

import pytest

import openpyxl
from io import BytesIO

from mbu_dev_shared_components.msoffice365.sharepoint_api.files import Sharepoint
# from mbu_dev_shared_components.msoffice365.excel.excel_reader import ExcelReader

FOLDER_NAME = "MSOffice tests"

CURRENT_DATE = datetime.datetime.now().strftime("%d-%m-%Y")
CURRENT_TIME = datetime.datetime.now().strftime("%H:%M")


@pytest.fixture(scope="module")
def sharepoint_api():
    """
    Fixture to authenticate with SharePoint and return a SharePoint API instance.

    This ensures all SharePoint-related operations across the test module use the same authenticated session.
    Skips the test suite if required environment variables are not set.
    """

    def _get_cfg(key: str) -> str:
        val = os.getenv(key)

        if not val:
            pytest.skip(f"env var '{key}' not set → skipping")

        return val

    username = _get_cfg("SvcRpaMBU002_USERNAME")
    password = _get_cfg("SvcRpaMBU002_PASSWORD")

    site_url = "https://aarhuskommune.sharepoint.com"
    site_name = "MBURPA"
    document_library = "Delte dokumenter"

    sp = Sharepoint(username=username, password=password, site_url=site_url, site_name=site_name, document_library=document_library)

    assert sp.ctx is not None, "Authentication failed: ClientContext is None"

    return sp


# ---------------------------------------------------------------------------
# TESTS
# ---------------------------------------------------------------------------
@pytest.mark.dependency(name="list_files")
def test_list_files_from_sharepoint_folder(sharepoint_api: Sharepoint):
    """
    Tests that a known Excel file exists in the target SharePoint folder.

    Verifies that SharePoint is reachable and that expected test files are present before running further tests.
    """

    files = sharepoint_api.fetch_files_list(FOLDER_NAME)

    assert isinstance(files, list), "Expected a list of files"

    file_names = [f["Name"] for f in files]

    assert "Test_Append_Rows.xlsx" in file_names, "Expected Test_Append_Rows.xlsx to be in list of files"


@pytest.mark.dependency(name="append_row", depends=["list_files"])
def test_append_row_to_excel(sharepoint_api: Sharepoint):
    """
    Tests that a new row can be appended to an existing Excel file hosted on SharePoint.

    Adds a row with the current date and time, then verifies that the row was successfully appended to the bottom of the sheet.
    """

    file_name = "Test_Append_Rows.xlsx"
    sheet_name = "Upload logs"

    headers = ["Upload date", "Upload time"]

    data = {
        "Upload date": CURRENT_DATE,
        "Upload time": CURRENT_TIME
    }

    sharepoint_api.append_row_to_sharepoint_excel(
        required_headers=headers,
        folder_name=FOLDER_NAME,
        excel_file_name=file_name,
        sheet_name=sheet_name,
        new_row=data,
    )

    binary_file = sharepoint_api.fetch_file_using_open_binary(file_name, FOLDER_NAME)

    wb = openpyxl.load_workbook(BytesIO(binary_file))

    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=True))

    newest_row = all_rows[-1]

    assert newest_row[0] == CURRENT_DATE

    assert newest_row[1] == CURRENT_TIME


@pytest.mark.dependency(depends=["append_row"])
def test_format_and_sort_excel_file(sharepoint_api: Sharepoint):
    """
    Tests that the Excel file is properly formatted and sorted after an update.

    This test fetches the file before and after formatting/sorting, and verifies that:
    - The top and bottom rows have changed due to sorting
    - The newest row (based on CURRENT_DATE) is now at the top (due to descending sort)
    """

    file_name = "Test_Append_Rows.xlsx"
    sheet_name = "Upload logs"

    test_file = sharepoint_api.fetch_file_using_open_binary(file_name=file_name, folder_name=FOLDER_NAME)

    wb = openpyxl.load_workbook(BytesIO(test_file))
    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=True))

    top_row_before_sorting = all_rows[1]
    bottom_row_before_sorting = all_rows[-1]

    sorting_keys = [
        {"key": "A", "ascending": False, "type": "str"},
        {"key": "B", "ascending": False, "type": "str"}
    ]

    sharepoint_api.format_and_sort_excel_file(
        folder_name=FOLDER_NAME,
        excel_file_name=file_name,
        sheet_name=sheet_name,
        sorting_keys=sorting_keys,
        bold_rows=[1],
        font_config=None,
        column_widths="auto",
        align_horizontal=None,
        align_vertical=None,
        freeze_panes=None,
    )

    test_file = sharepoint_api.fetch_file_using_open_binary(file_name=file_name, folder_name=FOLDER_NAME)

    wb = openpyxl.load_workbook(BytesIO(test_file))
    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=True))

    top_row_after_sorting = all_rows[1]
    bottom_row_after_sorting = all_rows[-1]

    assert top_row_before_sorting != top_row_after_sorting, "Expected top row to not be the same after sorting"

    assert bottom_row_before_sorting != bottom_row_after_sorting, "Expected bottom row to not be the same after sorting"

    assert top_row_after_sorting[0] == CURRENT_DATE
